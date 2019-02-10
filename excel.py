from openpyxl import Workbook
from openpyxl import load_workbook
import sys

def parameters():
	#function to break out the sys.argv array into global variables, enabling use of operators and filenames
	pass

def worksheets(workbook):
	worksheet1 = workbook.active
	worksheet1.title = "worksheet1"
	worksheet1.sheet_properties.tabColor = "1072BA"
	return worksheet1

def cells(worksheet):
	worksheet['D4'] = 436
	cell1 = worksheet['D4']
	print "Value in cell is:", cell1

def open_workbook (filename):
	print 'filename passed: ', filename
	try:
		workbook_name = load_workbook(filename)
	except IOError:
		new_file_prompt = str(raw_input('Create new workbook? Y/N : '))
		if new_file_prompt == 'Y' or 'y':
			workbook_name = new_workbook()
			worksheet = worksheets(workbook_name)
	return workbook_name, worksheet

def new_workbook():
		workbook_name = Workbook()
		return workbook_name

def print_worksheet_names(workbook):
	print 'The workbook has the following worksheets:'
	for sheet in workbook:
		print(sheet.title), '\n'

def save_workbook(workbook,filename):
	workbook.save(filename)
	print ('%s saved' % filename)

def change_cell():
	#modify the name and function of this feature with something useful.
	raw_input("You get to do something with the workbook here, press any key to continue: \n")

def main(arg_array):
	print 'Checking inputs...'
	print 'Number of arguments:', len(arg_array), 'arguments.'
	print 'Argument List:', str(arg_array), '\n'
	try:
		filename = str(arg_array[1])
		workbook, worksheet = open_workbook(filename)
	except IndexError:
		filename = str(raw_input('What do you want to call this file (remember to add the .xlsx extention : '))
		workbook = new_workbook()
		worksheet = worksheets(workbook)
	cells(worksheet)
	print_worksheet_names(workbook)
	change_cell()
	save_workbook(workbook,filename)

main(sys.argv)



